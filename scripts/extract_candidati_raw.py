import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
from collections import defaultdict

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
registry = {}
names = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc == 'TIMIȘOARA':
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc
        names[code] = row[idx['Denumire']]

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
MIN_N = 15

per_year = {}
for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(round(float(media), 2))
    wbx.close()
    per_year[year] = {code: vals for code, vals in acc.items() if len(vals) >= MIN_N}
    print(year, 'scoli:', len(per_year[year]), 'total candidati:', sum(len(v) for v in per_year[year].values()))

# ordinea fixa: rang shrink 2025 (din shrinkage_mediana.json), apoi ce lipseste din 2025 dupa ultimul an disponibil
shrink = json.load(open(f"{BASE}\\shrinkage_mediana.json", encoding='utf-8'))
order2025 = [s['cod'] for s in sorted(shrink['2025']['scoli'], key=lambda s: s['rang_shrink'])]
all_codes = set()
for year in YEARS:
    all_codes.update(per_year[year].keys())
missing = [c for c in all_codes if c not in order2025]
# sorteaza missing dupa ultimul rang disponibil in orice an anterior
def last_rank(code):
    for year in reversed(YEARS):
        ys = shrink.get(str(year), {}).get('scoli', [])
        for s in ys:
            if s['cod'] == code:
                return s['rang_shrink']
    return 999
missing.sort(key=last_rank)
fixed_order = order2025 + missing
print('total scoli in ordinea fixa:', len(fixed_order))

out = {
    'order': fixed_order,
    'names': {c: names[c] for c in fixed_order},
    'ani': {str(y): per_year[y] for y in YEARS},
}
op = f"{BASE}\\candidati_raw_timisoara.json"
with open(op, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False)
print('saved', op)
import os
print('size KB:', round(os.path.getsize(op) / 1024, 1))
