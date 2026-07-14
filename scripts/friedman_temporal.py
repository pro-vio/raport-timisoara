import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import math
from collections import defaultdict

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}
names = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc
        names[code] = row[idx['Denumire']]

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
MIN_N = 8

matrix = defaultdict(dict)  # code -> {year: (mean, n)}
for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(float(media))
    wbx.close()
    for code, vals in acc.items():
        if len(vals) >= MIN_N:
            matrix[code][year] = (round(sum(vals) / len(vals), 4), len(vals))

# save school x year matrix
mat_out = []
for code, ys in matrix.items():
    mat_out.append({
        'cod_siiir': code, 'denumire': names[code], 'localitate': registry[code],
        'medii_pe_ani': {str(y): ys[y][0] for y in YEARS if y in ys},
        'n_pe_ani': {str(y): ys[y][1] for y in YEARS if y in ys},
    })
with open(f"{BASE}\\medii_scoala_x_an.json", 'w', encoding='utf-8') as f:
    json.dump(mat_out, f, ensure_ascii=False, indent=1)

# chi-square upper tail via regularized upper incomplete gamma
def chi2_sf(x, df):
    a = df / 2.0
    x2 = x / 2.0
    if x2 < a + 1:
        # lower series
        term = 1.0 / a
        s = term
        n = 0
        while True:
            n += 1
            term *= x2 / (a + n)
            s += term
            if term < s * 1e-14:
                break
        lower = s * math.exp(-x2 + a * math.log(x2) - math.lgamma(a))
        return 1.0 - lower
    # upper continued fraction (Lentz)
    tiny = 1e-300
    b = x2 + 1 - a
    c = 1 / tiny
    d = 1 / b
    h = d
    for i in range(1, 300):
        an = -i * (i - a)
        b += 2
        d = an * d + b
        if abs(d) < tiny: d = tiny
        c = b + an / c
        if abs(c) < tiny: c = tiny
        d = 1 / d
        delta = d * c
        h *= delta
        if abs(delta - 1) < 1e-14:
            break
    return h * math.exp(-x2 + a * math.log(x2) - math.lgamma(a))

def rank_row(vals):
    order = sorted(range(len(vals)), key=lambda i: vals[i])
    ranks = [0.0] * len(vals)
    i = 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and vals[order[j + 1]] == vals[order[i]]:
            j += 1
        avg = (i + j) / 2 + 1
        for kk in range(i, j + 1):
            ranks[order[kk]] = avg
        i = j + 1
    return ranks

CITIES = ['CLUJ-NAPOCA', 'IAȘI', 'TIMIȘOARA']
k = len(YEARS)
result = {}
print(f'Friedman per oraș: blocuri = școli prezente în toți cei {k} ani (min {MIN_N} candidați/an); tratament = anul')
for city in CITIES:
    blocks = []
    for code, ys in matrix.items():
        if registry[code] == city and all(y in ys for y in YEARS):
            blocks.append([ys[y][0] for y in YEARS])
    n = len(blocks)
    rank_sums = [0.0] * k
    tie_corr_num = 0.0
    for b in blocks:
        rr = rank_row(b)
        for i, r in enumerate(rr):
            rank_sums[i] += r
        # tie correction per block
        from collections import Counter
        cnt = Counter(b)
        tie_corr_num += sum(t ** 3 - t for t in cnt.values())
    Q = 12.0 / (n * k * (k + 1)) * sum(rs * rs for rs in rank_sums) - 3 * n * (k + 1)
    denom = 1 - tie_corr_num / (n * k * (k * k - 1))
    if denom > 0:
        Q /= denom
    p = chi2_sf(Q, k - 1)
    W = Q / (n * (k - 1))  # Kendall's W
    mean_ranks = [round(rs / n, 2) for rs in rank_sums]
    result[city] = {'n_scoli_balansate': n, 'Q': round(Q, 3), 'df': k - 1,
                    'p': float(f'{p:.4g}'), 'kendall_W': round(W, 4),
                    'rang_mediu_pe_an': dict(zip(map(str, YEARS), mean_ranks))}
    print(f'\n{city}: n={n} școli, Q={Q:.2f}, df={k-1}, p={p:.4g}, W={W:.3f}')
    print('  rang mediu pe an (1=cel mai slab an, 6=cel mai bun):', dict(zip(YEARS, mean_ranks)))

with open(f"{BASE}\\friedman_pe_orase.json", 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print('\nsaved medii_scoala_x_an.json + friedman_pe_orase.json')
