# Inventar de structură pentru fișierele BAC sesiunea de vară 2016-2025 (date/bac/):
# antet (lista coloanelor), un rând de exemplu, număr de rânduri. Scop: verificarea
# identificatorului de școală (SIIIR? denumire?) și a schimbărilor de schemă între ani.
import sys, os, csv, glob
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

BASE = os.path.join(os.path.dirname(__file__), '..', 'date', 'bac')

def head_csv(path):
    # încearcă mai multe combinații encoding/delimitator
    for enc in ('utf-8-sig', 'utf-8', 'cp1250', 'latin-1'):
        try:
            with open(path, encoding=enc, newline='') as f:
                sample = f.read(65536)
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                f.seek(0)
                rd = csv.reader(f, dialect)
                header = next(rd)
                first = next(rd)
                n = 1 + sum(1 for _ in rd)
            return enc, dialect.delimiter, header, first, n
        except (UnicodeDecodeError, csv.Error):
            continue
    raise RuntimeError(f'nu pot citi {path}')

def head_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    first = list(next(rows))
    return wb.sheetnames, ws.max_row, header, first

for path in sorted(glob.glob(os.path.join(BASE, 'bac_*'))):
    name = os.path.basename(path)
    print('=' * 100)
    print(name)
    try:
        if name.endswith('.csv'):
            enc, delim, header, first, n = head_csv(path)
            print(f'  CSV encoding={enc} delim={delim!r} rânduri_date={n:,}')
        else:
            sheets, maxrow, header, first = head_xlsx(path)
            print(f'  XLSX sheets={sheets} max_row(metadata)={maxrow:,}')
        print(f'  coloane ({len(header)}):')
        for i, (h, v) in enumerate(zip(header, first)):
            print(f'    [{i:2d}] {h!r:60s} ex: {str(v)[:45]!r}')
    except Exception as e:
        print(f'  EROARE: {type(e).__name__}: {e}')
