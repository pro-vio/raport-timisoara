import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import numpy as np
from collections import defaultdict

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}

targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = norm(row[idx['Localitate']])

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"

by_city = defaultdict(list)
for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        city = registry.get(code)
        if city is None:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            by_city[city].append(float(media))
    wbx.close()

bins = np.arange(1.0, 10.25, 0.25)
out = {}
for city, vals in by_city.items():
    hist, edges = np.histogram(vals, bins=bins)
    out[city] = {
        'n': len(vals),
        'edges': [round(float(e), 2) for e in edges],
        'counts': [int(c) for c in hist],
    }

out_path = f"{BASE}\\histograme_orase.json"
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False)
print('saved', out_path)
for c, d in out.items():
    print(c, d['n'])
