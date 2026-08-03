# Elevii înscriși care nu se prezintă la examen: cât cântăresc, unde se strâng, și dacă
# poziția Timișoarei ține de ei.
#
# Verificarea cerută de user pe 3 august 2026, ca să nu rămână o explicație plauzibilă
# nescrisă: corelația dintre nivelul școlii și ponderea neprezentaților la ea s-ar putea
# să fie circulară — neprezentații trag mediana în jos, deci școlile cu mulți neprezentați
# ar avea mediana mică prin construcție. De aceea nivelul școlii se citește sub trei măsuri:
#   A. mediana notelor celor prezenți   (neprezentații NU intră în ea)
#   B. media notelor celor prezenți     (altă statistică, aceleași note)
#   C. mediana lui MEDIA V-VIII         (notele claselor V-VIII, din afara examenului)
# Dacă cele trei dau același semn și aceeași mărime, corelația nu e un artefact.
import io, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import xlrd
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from statistici import median_of, mediana_cu_neprezentati
from praguri import prag

DATE = os.path.join(HERE, '..', 'date')
ORASE = ['CLUJ-NAPOCA', 'IAȘI', 'TIMIȘOARA']
YEARS = list(range(2020, 2026))
MIN_N = prag('prag_teste')      # calculat, nu ales; vezi prag_incertitudine.py
OUT = os.path.join(DATE, 'neprezentati.json')


def norm(s):
    if not isinstance(s, str):
        return s
    return (s.strip().upper().replace('Ş', 'Ș').replace('ş', 'ș')
             .replace('Ţ', 'Ț').replace('ţ', 'ț'))


def cod(v):
    c = str(v).strip()
    return c[:-2] if c.endswith('.0') else c


def pearson(xs, ys):
    n = len(xs)
    mx, my = sum(xs) / n, sum(ys) / n
    sx = (sum((x - mx) ** 2 for x in xs) / n) ** .5
    sy = (sum((y - my) ** 2 for y in ys) / n) ** .5
    return None if sx == 0 or sy == 0 else \
        sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / (n * sx * sy)


wb = xlrd.open_workbook(os.path.join(DATE, 'Unitati de invatamant acreditate  i autorizate.xls'))
ws = wb.sheet_by_index(0)
idx = {h: i for i, h in enumerate(ws.row_values(0))}
reg = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    if norm(row[idx['Localitate']]) in set(ORASE):
        reg[cod(row[idx['Cod']])] = norm(row[idx['Localitate']])

randuri = []
pondere = defaultdict(dict)
for year in YEARS:
    wbx = openpyxl.load_workbook(os.path.join(DATE, f'{year}_evnat_date-deschise.xlsx'),
                                 read_only=(year != 2020))
    wsx = wbx[wbx.sheetnames[0]]
    it = wsx.iter_rows(min_row=1, values_only=True)
    hd = list(next(it))
    hi = {h.strip(): i for i, h in enumerate(hd) if isinstance(h, str)}
    i_cod, i_med, i_viii = hi['COD SIIIR'], hi['MEDIA'], hi['MEDIA V-VIII']
    note, viii, fara = defaultdict(list), defaultdict(list), defaultdict(int)
    for row in it:
        if row[i_cod] is None:
            continue
        c = cod(row[i_cod])
        if c not in reg:
            continue
        if isinstance(row[i_med], (int, float)):
            note[c].append(float(row[i_med]))
        else:
            fara[c] += 1
        if isinstance(row[i_viii], (int, float)):
            viii[c].append(float(row[i_viii]))
    wbx.close()

    for oras in ORASE:
        nn = sum(len(v) for c, v in note.items() if reg[c] == oras)
        na = sum(n for c, n in fara.items() if reg[c] == oras)
        pondere[oras][str(year)] = round(100 * na / (nn + na), 2)

    for c, v in note.items():
        if len(v) >= MIN_N:
            na = fara.get(c, 0)
            randuri.append({'an': year, 'oras': reg[c],
                            'mediana_en': median_of(v), 'media_en': sum(v) / len(v),
                            'mediana_viii': median_of(viii[c]) if viii[c] else None,
                            'pond_neprez': na / (len(v) + na),
                            'mediana_cu': mediana_cu_neprezentati(v, na)})
    print(f'{year}: {sum(1 for r in randuri if r["an"] == year)} celule școală-an')

cor = {}
for oras in ORASE + ['TOATE']:
    sel = [r for r in randuri if oras == 'TOATE' or r['oras'] == oras]
    s3 = [r for r in sel if r['mediana_viii'] is not None]
    cor[oras] = {
        'mediana_en': round(pearson([r['mediana_en'] for r in sel],
                                    [r['pond_neprez'] for r in sel]), 3),
        'media_en': round(pearson([r['media_en'] for r in sel],
                                  [r['pond_neprez'] for r in sel]), 3),
        'mediana_viii': round(pearson([r['mediana_viii'] for r in s3],
                                      [r['pond_neprez'] for r in s3]), 3),
        'n': len(sel),
    }

# cât se mișcă mediana medianelor unui oraș când neprezentații intră în calcul
miscare = {}
for oras in ORASE:
    miscare[oras] = {}
    for year in YEARS:
        sel = [r for r in randuri if r['oras'] == oras and r['an'] == year]
        fara_ = median_of([r['mediana_en'] for r in sel])
        cu_ = median_of([r['mediana_cu'] for r in sel if r['mediana_cu'] is not None])
        miscare[oras][str(year)] = round(cu_ - fara_, 3)

out = {'min_candidati_per_scoala_an': MIN_N,
       'pondere_neprezentati_pct': pondere,
       'corelatie_nivel_vs_pondere': cor,
       'miscarea_medianei_orasului': miscare}
io.open(OUT, 'w', encoding='utf-8', newline='\n').write(
    json.dumps(out, ensure_ascii=False, indent=1))
print()
print('corelații nivel × pondere neprezentați:')
for o, v in cor.items():
    print(f"  {o:12s} mediana EN {v['mediana_en']:+.3f}  media EN {v['media_en']:+.3f}  "
          f"mediana V-VIII {v['mediana_viii']:+.3f}  (n={v['n']})")
print('salvat:', os.path.normpath(OUT))
