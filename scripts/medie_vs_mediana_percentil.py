import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
from collections import defaultdict

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
registry = {}
names = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc == 'TIMIȘOARA':
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc
        names[code] = row[idx['Denumire']]

print('scoli Timisoara in registru:', len(registry))

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
MIN_N = 15

def pct_of_mean(values, mean):
    below = sum(1 for v in values if v < mean)
    equal = sum(1 for v in values if v == mean)
    n = len(values)
    return 100.0 * (below + 0.5 * equal) / n

def median_of(values):
    v = sorted(values)
    n = len(v)
    return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2

per_year_results = {}
all_school_year_rows = []

for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(float(media))
    wbx.close()

    gaps = []
    for code, vals in acc.items():
        if len(vals) < MIN_N:
            continue
        mean = sum(vals) / len(vals)
        med = median_of(vals)
        pct = pct_of_mean(vals, mean)
        gap = pct - 50.0
        gaps.append(gap)
        all_school_year_rows.append({
            'an': year, 'cod_siiir': code, 'denumire': names[code],
            'n': len(vals), 'medie': round(mean, 3), 'mediana': round(med, 3),
            'percentila_mediei': round(pct, 1), 'gap_percentile': round(gap, 1),
        })

    gaps_sorted = sorted(gaps)
    ng = len(gaps_sorted)
    med_gap = gaps_sorted[ng // 2] if ng % 2 else (gaps_sorted[ng // 2 - 1] + gaps_sorted[ng // 2]) / 2
    avg_gap = sum(gaps) / ng
    per_year_results[year] = {
        'n_scoli': ng, 'gap_mediu': round(avg_gap, 2), 'gap_median': round(med_gap, 2),
        'gap_min': round(min(gaps), 2), 'gap_max': round(max(gaps), 2),
    }
    print(f"{year}: n_scoli={ng}  gap mediu={avg_gap:.2f}pp  gap median={med_gap:.2f}pp  min={min(gaps):.2f}  max={max(gaps):.2f}")

# corelatie gap vs nivelul mediei scolii (pooled peste toti anii)
xs = [r['medie'] for r in all_school_year_rows]
ys = [r['gap_percentile'] for r in all_school_year_rows]
n = len(xs)
mx = sum(xs) / n
my = sum(ys) / n
cov = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / n
sx = (sum((x - mx) ** 2 for x in xs) / n) ** 0.5
sy = (sum((y - my) ** 2 for y in ys) / n) ** 0.5
corr = cov / (sx * sy)
print(f"\ncorelatie (medie scolii) vs (gap percentila mediei fata de mediana): r={corr:.3f}  n={n} randuri scoala-an")

all_school_year_rows.sort(key=lambda r: r['gap_percentile'])
print("\ncele mai negative gap-uri (media MULT sub percentila 50, coada stanga lunga -> scoli de elita):")
for r in all_school_year_rows[:8]:
    print(' ', r['an'], r['denumire'], 'n=', r['n'], 'medie=', r['medie'], 'mediana=', r['mediana'], 'percentila mediei=', r['percentila_mediei'])
print("\ncele mai pozitive gap-uri (media peste percentila 50, coada dreapta / scoli slabe):")
for r in all_school_year_rows[-8:]:
    print(' ', r['an'], r['denumire'], 'n=', r['n'], 'medie=', r['medie'], 'mediana=', r['mediana'], 'percentila mediei=', r['percentila_mediei'])

out = {'per_an': per_year_results, 'corelatie_medie_vs_gap': round(corr, 3), 'scoala_an': all_school_year_rows}
op = f"{BASE}\\medie_vs_mediana_percentil.json"
with open(op, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print('\nsaved', op)
