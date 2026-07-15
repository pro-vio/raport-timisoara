import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import math
import gc
from collections import defaultdict, Counter

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

BASE = r"C:\Users\Viorel Proteasa\Documents\raport-timisoara\date"
REG = BASE + r"\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}
names = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc
        names[code] = row[idx['Denumire']]

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
MIN_N = 8

def median_of(values):
    v = sorted(values)
    n = len(v)
    return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2

matrix = defaultdict(dict)  # code -> {year: median}
for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(float(media))
    wbx.close()
    for code, vals in acc.items():
        if len(vals) >= MIN_N:
            matrix[code][year] = round(median_of(vals), 4)
    del wbx, wsx, rows_iter, acc
    gc.collect()
    print(f'{year}: done')

def rank_row(vals):
    order = sorted(range(len(vals)), key=lambda i: vals[i])
    ranks = [0.0] * len(vals)
    i = 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and vals[order[j + 1]] == vals[order[i]]:
            j += 1
        avg = (i + j) / 2 + 1
        for kk in range(i, j + 1):
            ranks[order[kk]] = avg
        i = j + 1
    return ranks

def chi2_sf(x, df):
    a = df / 2.0
    x2 = x / 2.0
    if x2 < a + 1:
        term = 1.0 / a
        s = term
        n = 0
        while True:
            n += 1
            term *= x2 / (a + n)
            s += term
            if term < s * 1e-14:
                break
        lower = s * math.exp(-x2 + a * math.log(x2) - math.lgamma(a))
        return 1.0 - lower
    tiny = 1e-300
    b = x2 + 1 - a
    c = 1 / tiny
    d = 1 / b
    h = d
    for i in range(1, 300):
        an = -i * (i - a)
        b += 2
        d = an * d + b
        if abs(d) < tiny: d = tiny
        c = b + an / c
        if abs(c) < tiny: c = tiny
        d = 1 / d
        delta = d * c
        h *= delta
        if abs(delta - 1) < 1e-14:
            break
    return h * math.exp(-x2 + a * math.log(x2) - math.lgamma(a))

CITIES = ['CLUJ-NAPOCA', 'IAȘI', 'TIMIȘOARA']
k = len(YEARS)
result = {}
for city in CITIES:
    blocks = []
    codes = []
    for code, ys in matrix.items():
        if registry[code] == city and all(y in ys for y in YEARS):
            blocks.append([ys[y] for y in YEARS])
            codes.append(code)
    n = len(blocks)
    rank_sums = [0.0] * k
    per_year_ranks = [[] for _ in range(k)]
    tie_corr_num = 0.0
    all_school_ranks = {}
    for b, code in zip(blocks, codes):
        rr = rank_row(b)
        all_school_ranks[code] = rr
        for i, r in enumerate(rr):
            rank_sums[i] += r
            per_year_ranks[i].append(r)
        cnt = Counter(b)
        tie_corr_num += sum(t ** 3 - t for t in cnt.values())
    Q = 12.0 / (n * k * (k + 1)) * sum(rs * rs for rs in rank_sums) - 3 * n * (k + 1)
    denom = 1 - tie_corr_num / (n * k * (k * k - 1))
    if denom > 0:
        Q /= denom
    p = chi2_sf(Q, k - 1)
    W = Q / (n * (k - 1))
    mean_ranks = [round(rs / n, 2) for rs in rank_sums]
    min_ranks = [round(min(pr), 2) for pr in per_year_ranks]
    max_ranks = [round(max(pr), 2) for pr in per_year_ranks]
    result[city] = {
        'n_scoli': n, 'Q': round(Q, 3), 'df': k - 1, 'p': float(f'{p:.4g}'),
        'kendall_W': round(W, 4),
        'rang_mediu': dict(zip(map(str, YEARS), mean_ranks)),
        'rang_min': dict(zip(map(str, YEARS), min_ranks)),
        'rang_max': dict(zip(map(str, YEARS), max_ranks)),
    }
    print(f'{city}: n={n}, Q={Q:.2f}, p={p:.4g}, W={W:.3f}')
    print('  rang mediu:', mean_ranks)
    print('  rang min:  ', min_ranks)
    print('  rang max:  ', max_ranks)

# exemplu pt textul explicativ: o școală mare din Timișoara cu mediane distincte
example_code = None
for code, ys in matrix.items():
    if registry[code] == 'TIMIȘOARA' and all(y in ys for y in YEARS):
        vals = [ys[y] for y in YEARS]
        if len(set(vals)) == 6:  # fara egaluri, exemplu curat
            if names[code].startswith('ȘCOALA GIMNAZIALĂ NR.16') or names[code].startswith('LICEUL TEORETIC "GRIGORE'):
                example_code = code
                break
if example_code is None:
    for code, ys in matrix.items():
        if registry[code] == 'TIMIȘOARA' and all(y in ys for y in YEARS) and len(set(ys[y] for y in YEARS)) == 6:
            example_code = code
            break
ex = {'denumire': names[example_code],
      'mediane': {str(y): matrix[example_code][y] for y in YEARS},
      'ranguri': dict(zip(map(str, YEARS), rank_row([matrix[example_code][y] for y in YEARS])))}
print('\nexemplu:', ex['denumire'])
print(' mediane:', ex['mediane'])
print(' ranguri:', ex['ranguri'])
result['_exemplu'] = ex

op = BASE + r"\friedman_mediane.json"
with open(op, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print('\nsaved', op)
