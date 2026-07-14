import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')  # Ş/ş -> Ș/ș
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')  # Ţ/ţ -> Ț/ț
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}

targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}  # code(str) -> dict
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        # xlrd may give float for numeric-looking codes
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = {
            'denumire': row[idx['Denumire']],
            'localitate': row[idx['Localitate']],
            'judet': row[idx['Județ']],
            'tip': row[idx['Tip unitate']],
        }

print('registry candidate codes in target cities:', len(registry))

EN_FILE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\2025_evnat_date-deschise.xlsx"
wbx = openpyxl.load_workbook(EN_FILE, read_only=True)
wsx = wbx[wbx.sheetnames[0]]
header_row = [c.value for c in next(wsx.iter_rows(min_row=1, max_row=1))]
hidx = {h.strip() if isinstance(h, str) else h: i for i, h in enumerate(header_row)}
siiir_col = hidx.get('COD SIIIR') if 'COD SIIIR' in hidx else hidx.get(' COD SIIIR')
media_col = hidx.get('MEDIA')

from collections import defaultdict
counts = defaultdict(int)
sums = defaultdict(float)
for row in wsx.iter_rows(min_row=2, values_only=True):
    code = row[siiir_col]
    if code is None:
        continue
    code = str(code).strip()
    if code.endswith('.0'):
        code = code[:-2]
    if code in registry:
        counts[code] += 1
        media = row[media_col]
        if isinstance(media, (int, float)):
            sums[code] += media
wbx.close()

print('matched codes with EN 2025 candidates:', len(counts))

result = []
for code, info in registry.items():
    if code in counts:
        n = counts[code]
        avg = round(sums[code] / n, 2) if n else None
        result.append({
            'cod_siiir': code,
            'denumire': info['denumire'],
            'localitate': info['localitate'],
            'judet': info['judet'],
            'nr_candidati_2025': n,
            'media_2025': avg,
        })

result.sort(key=lambda x: (x['localitate'], -x['nr_candidati_2025']))
print('total scoli identificate (cu candidati EN 2025):', len(result))

out_path = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\siiir_timisoara_cluj_iasi.json"
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print('saved to', out_path)

for city in ['TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI']:
    rows = [r for r in result if r['localitate'].upper() == city or norm(r['localitate']) == city]
    print(f'--- {city}: {len(rows)} scoli ---')
    for r in rows[:10]:
        print(r['cod_siiir'], r['denumire'], r['nr_candidati_2025'], r['media_2025'])
