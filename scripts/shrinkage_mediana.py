import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import random
from collections import defaultdict

random.seed(20260707)

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
registry = {}
names = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc == 'TIMIȘOARA':
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc
        names[code] = row[idx['Denumire']]

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
MIN_N = 15
B = 2000

def median_of(values):
    v = sorted(values)
    n = len(v)
    return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2

def bootstrap_se_median(values, B):
    n = len(values)
    meds = []
    for _ in range(B):
        sample = [values[random.randrange(n)] for _ in range(n)]
        meds.append(median_of(sample))
    mean_b = sum(meds) / B
    var_b = sum((m - mean_b) ** 2 for m in meds) / (B - 1)
    return var_b ** 0.5

results_per_year = {}

for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(float(media))
    wbx.close()

    schools = []
    for code, vals in acc.items():
        if len(vals) < MIN_N:
            continue
        m = median_of(vals)
        se = bootstrap_se_median(vals, B)
        schools.append({'cod': code, 'denumire': names[code], 'n': len(vals), 'mediana': m, 'se': se})

    k = len(schools)
    ms = [s['mediana'] for s in schools]
    ses = [s['se'] for s in schools]
    mu_hat = sum(ms) / k
    var_ms = sum((m - mu_hat) ** 2 for m in ms) / (k - 1)
    mean_se2 = sum(se ** 2 for se in ses) / k
    tau2 = max(0.0, var_ms - mean_se2)

    for s in schools:
        se2 = s['se'] ** 2
        if tau2 <= 1e-9:
            w = 0.0
        else:
            w = tau2 / (tau2 + se2)
        theta = w * s['mediana'] + (1 - w) * mu_hat
        if tau2 <= 1e-9:
            post_var = se2
        else:
            post_var = 1.0 / (1.0 / tau2 + 1.0 / se2)
        s['w_shrink'] = round(w, 3)
        s['mediana_shrink'] = round(theta, 3)
        s['se_shrink'] = round(post_var ** 0.5, 3)
        s['ci_low'] = round(theta - 1.96 * post_var ** 0.5, 3)
        s['ci_high'] = round(theta + 1.96 * post_var ** 0.5, 3)
        s['mediana'] = round(s['mediana'], 3)
        s['se'] = round(s['se'], 3)

    schools_naiv = sorted(schools, key=lambda s: -s['mediana'])
    for i, s in enumerate(schools_naiv):
        s['rang_naiv'] = i + 1
    schools_shrink = sorted(schools, key=lambda s: -s['mediana_shrink'])
    for i, s in enumerate(schools_shrink):
        s['rang_shrink'] = i + 1

    results_per_year[year] = {
        'mu_hat': round(mu_hat, 3), 'tau2': round(tau2, 4), 'n_scoli': k,
        'scoli': sorted(schools, key=lambda s: s['rang_shrink']),
    }
    print(f"{year}: n_scoli={k}  mu_hat={mu_hat:.3f}  tau2={tau2:.4f}")
    max_shift = max(abs(s['rang_naiv'] - s['rang_shrink']) for s in schools)
    print(f"  cea mai mare schimbare de rang (naiv vs shrink): {max_shift}")

op = f"{BASE}\\shrinkage_mediana.json"
with open(op, 'w', encoding='utf-8') as f:
    json.dump(results_per_year, f, ensure_ascii=False, indent=1)
print('saved', op)

print("\n=== 2025: clasament complet (shrink) ===")
for s in results_per_year[2025]['scoli']:
    print(f"  {s['rang_shrink']:2d}. {s['denumire']:<55s} n={s['n']:4d}  mediana_bruta={s['mediana']:.2f} (rang naiv {s['rang_naiv']:2d})  shrink={s['mediana_shrink']:.2f} [{s['ci_low']:.2f}, {s['ci_high']:.2f}]  w={s['w_shrink']}")
