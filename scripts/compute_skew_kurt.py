import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import numpy as np
from collections import defaultdict

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}

targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = {
            'denumire': row[idx['Denumire']],
            'localitate': row[idx['Localitate']],
            'judet': row[idx['Județ']],
        }

print('registry candidate codes in target cities:', len(registry))

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
FILES = {
    2020: 'evnat_date-deschise_2020.xlsx' if False else '2020_evnat_date-deschise.xlsx',
    2021: '2021_evnat_date-deschise.xlsx',
    2022: '2022_evnat_date-deschise.xlsx',
    2023: '2023_evnat_date-deschise.xlsx',
    2024: '2024_evnat_date-deschise.xlsx',
    2025: '2025_evnat_date-deschise.xlsx',
}

by_school = defaultdict(list)   # code -> list of (year, media)
by_city = defaultdict(list)     # localitate -> list of media

for year in YEARS:
    path = f"{BASE}\\{FILES[year]}"
    ro = year != 2020  # 2020 has broken dimension metadata in read_only mode
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {}
    for i, h in enumerate(header_row):
        if isinstance(h, str):
            hidx[h.strip()] = i
    siiir_col = hidx.get('COD SIIIR')
    media_col = hidx.get('MEDIA')
    n_rows = 0
    n_matched = 0
    for row in rows_iter:
        n_rows += 1
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        if code not in registry:
            continue
        media = row[media_col]
        if not isinstance(media, (int, float)):
            continue
        n_matched += 1
        by_school[code].append(media)
        by_city[registry[code]['localitate']].append(media)
    wbx.close()
    print(year, 'total candidates:', n_rows, 'matched (3 cities):', n_matched)

def skew_kurt(values):
    x = np.array(values, dtype=float)
    n = len(x)
    if n < 8:
        return None
    mean = x.mean()
    s = x.std(ddof=1)
    if s == 0:
        return {'n': n, 'mean': round(mean, 3), 'std': 0.0, 'skew': 0.0, 'kurt_excess': 0.0}
    z = (x - mean) / s
    g1 = (n / ((n - 1) * (n - 2))) * np.sum(z ** 3)
    g2 = (n * (n + 1) / ((n - 1) * (n - 2) * (n - 3))) * np.sum(z ** 4) - (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    return {
        'n': n,
        'mean': round(float(mean), 3),
        'std': round(float(s), 3),
        'skew': round(float(g1), 3),
        'kurt_excess': round(float(g2), 3),
    }

school_stats = []
for code, vals in by_school.items():
    stat = skew_kurt(vals)
    if stat is None:
        continue
    info = registry[code]
    school_stats.append({
        'cod_siiir': code,
        'denumire': info['denumire'],
        'localitate': info['localitate'],
        'judet': info['judet'],
        **stat,
    })

city_stats = {}
for city, vals in by_city.items():
    city_stats[city] = skew_kurt(vals)

school_stats.sort(key=lambda r: (r['localitate'], -r['n']))

out = {
    'schools': school_stats,
    'cities': city_stats,
}
out_path = f"{BASE}\\skew_kurt_2020_2025.json"
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print('n schools with stats:', len(school_stats))
print('city stats:')
for c, s in city_stats.items():
    print(' ', c, s)

print('\ntop 5 most negative skew (schools):')
for r in sorted(school_stats, key=lambda r: r['skew'])[:5]:
    print(' ', r['denumire'], r['localitate'], r['n'], r['skew'], r['kurt_excess'])
print('\ntop 5 most positive skew (schools):')
for r in sorted(school_stats, key=lambda r: -r['skew'])[:5]:
    print(' ', r['denumire'], r['localitate'], r['n'], r['skew'], r['kurt_excess'])
print('\ntop 5 highest excess kurtosis:')
for r in sorted(school_stats, key=lambda r: -r['kurt_excess'])[:5]:
    print(' ', r['denumire'], r['localitate'], r['n'], r['skew'], r['kurt_excess'])
print('\ntop 5 lowest (most negative) excess kurtosis:')
for r in sorted(school_stats, key=lambda r: r['kurt_excess'])[:5]:
    print(' ', r['denumire'], r['localitate'], r['n'], r['skew'], r['kurt_excess'])

print('saved to', out_path)
