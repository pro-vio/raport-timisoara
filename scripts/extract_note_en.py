# Extracția comună pentru lanțul Evaluare Națională: notele per școală×an, plus
# NEPREZENTAȚII, codați explicit.
#
# Până pe 3 august 2026 fiecare dintre cele cinci scripturi EN citea singur cele șase
# xlsx-uri, cu același bloc copiat, și arunca tăcut candidații fără medie. Aici se citește
# o singură dată, iar absența capătă nume.
#
# Trei stări ale lumii care înainte ajungeau în același gol:
#   * școala nu apare deloc în fișierul acelui an        -> lipsește cheia
#   * școala apare, dar candidații ei nu s-au prezentat  -> `neprezentati` > 0
#   * codul din fișier nu e în registrul SIIIR           -> numărat la `coduri_straine`
import io, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DATE = os.path.join(HERE, '..', 'date')
ORASE = {'CLUJ-NAPOCA', 'IAȘI', 'TIMIȘOARA'}
YEARS = list(range(2020, 2026))
OUT = os.path.join(DATE, 'note_en.json')


def norm(s):
    if not isinstance(s, str):
        return s
    return (s.strip().upper().replace('Ş', 'Ș').replace('ş', 'ș')
             .replace('Ţ', 'Ț').replace('ţ', 'ț'))


def cod(v):
    c = str(v).strip()
    return c[:-2] if c.endswith('.0') else c


wb = xlrd.open_workbook(os.path.join(DATE, 'Unitati de invatamant acreditate  i autorizate.xls'))
ws = wb.sheet_by_index(0)
idx = {h: i for i, h in enumerate(ws.row_values(0))}
registry, denumiri = {}, {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in ORASE:
        registry[cod(row[idx['Cod']])] = loc
        denumiri[cod(row[idx['Cod']])] = norm(row[idx['Denumire']])

out = {'orase': registry, 'denumiri': denumiri, 'ani': {}}
for year in YEARS:
    wbx = openpyxl.load_workbook(os.path.join(DATE, f'{year}_evnat_date-deschise.xlsx'),
                                 read_only=(year != 2020))
    wsx = wbx[wbx.sheetnames[0]]
    it = wsx.iter_rows(min_row=1, values_only=True)
    hd = list(next(it))
    hi = {h.strip(): i for i, h in enumerate(hd) if isinstance(h, str)}
    i_cod, i_med, i_v8 = hi['COD SIIIR'], hi['MEDIA'], hi['MEDIA V-VIII']
    note = defaultdict(list)
    # Perechea (media dată de școală în clasele V-VIII, media de la examen), PER ELEV.
    # Se păstrează aici fiindcă doar aici există împerecherea — lista de note se sortează
    # și o pierde. Din perechi se obțin și decalajul, și cât de bine ordonează școala.
    perechi = defaultdict(list)
    fara = defaultdict(int)
    straine = fara_cod = 0
    for row in it:
        if row[i_cod] is None:
            fara_cod += 1
            continue
        c = cod(row[i_cod])
        if c not in registry:
            straine += 1
            continue
        m = row[i_med]
        if isinstance(m, (int, float)):
            note[c].append(round(float(m), 2))
            if isinstance(row[i_v8], (int, float)):
                perechi[c].append([round(float(row[i_v8]), 2), round(float(m), 2)])
        else:
            fara[c] += 1          # înscris, fără medie: neprezentat sau eliminat
    wbx.close()

    scoli = {c: {'note': sorted(v), 'neprezentati': fara.get(c, 0),
                 'v8_en': perechi.get(c, [])} for c, v in note.items()}
    # școli din care s-au înscris doar neprezentați: există în an, cu zero note
    for c, n in fara.items():
        scoli.setdefault(c, {'note': [], 'neprezentati': n, 'v8_en': []})
    out['ani'][str(year)] = {'scoli': scoli, 'coduri_straine': straine,
                             'candidati_fara_cod': fara_cod}
    np = sum(s['neprezentati'] for s in scoli.values())
    nn = sum(len(s['note']) for s in scoli.values())
    print(f'{year}: {len(scoli)} școli · {nn} cu medie · {np} neprezentați '
          f'({100 * np / (nn + np):.1f}%) · {straine} coduri din afara celor trei orașe')

io.open(OUT, 'w', encoding='utf-8', newline='\n').write(
    json.dumps(out, ensure_ascii=False))
print('salvat:', os.path.normpath(OUT), f'({os.path.getsize(OUT) / 1e6:.1f} MB)')
