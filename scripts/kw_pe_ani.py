import sys
sys.stdout.reconfigure(encoding='utf-8')
import xlrd
import openpyxl
import json
import math
from collections import defaultdict, Counter

def norm(s):
    if not isinstance(s, str):
        return s
    s = s.strip().upper()
    s = s.replace('Ş', 'Ș').replace('ş', 'ș')
    s = s.replace('Ţ', 'Ț').replace('ţ', 'ț')
    return s

REG = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date\Unitati de invatamant acreditate  i autorizate.xls"
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
header = ws.row_values(0)
idx = {h: i for i, h in enumerate(header)}
targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
registry = {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    loc = norm(row[idx['Localitate']])
    if loc in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        registry[code] = loc

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
BASE = r"C:\Users\Viorel Proteasa\Documents\evaluare-nationala\date"
MIN_N = 8

# per year: code -> list of medii
per_year = {}
for year in YEARS:
    path = f"{BASE}\\{year}_evnat_date-deschise.xlsx"
    ro = year != 2020
    wbx = openpyxl.load_workbook(path, read_only=ro)
    wsx = wbx[wbx.sheetnames[0]]
    rows_iter = wsx.iter_rows(min_row=1, values_only=True)
    header_row = list(next(rows_iter))
    hidx = {h.strip(): i for i, h in enumerate(header_row) if isinstance(h, str)}
    siiir_col = hidx['COD SIIIR']
    media_col = hidx['MEDIA']
    acc = defaultdict(list)
    for row in rows_iter:
        code = row[siiir_col]
        if code is None:
            continue
        code = str(code).strip()
        if code.endswith('.0'):
            code = code[:-2]
        city = registry.get(code)
        if city is None:
            continue
        media = row[media_col]
        if isinstance(media, (int, float)):
            acc[code].append(float(media))
    wbx.close()
    per_year[year] = acc

def rank_all(values):
    order = sorted(range(len(values)), key=lambda i: values[i])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and values[order[j + 1]] == values[order[i]]:
            j += 1
        avg = (i + j) / 2 + 1
        for kk in range(i, j + 1):
            ranks[order[kk]] = avg
        i = j + 1
    return ranks

def kruskal(glist):
    all_v = [v for g in glist for v in g]
    N = len(all_v)
    ranks = rank_all(all_v)
    H = 0.0
    pos = 0
    rank_sums = []
    for g in glist:
        rs = sum(ranks[pos:pos + len(g)])
        rank_sums.append(rs)
        H += rs * rs / len(g)
        pos += len(g)
    H = 12.0 / (N * (N + 1)) * H - 3 * (N + 1)
    cnt = Counter(all_v)
    T = sum(t ** 3 - t for t in cnt.values())
    corr = 1 - T / (N ** 3 - N)
    if corr > 0:
        H /= corr
    return H, N, rank_sums

def norm_sf(z):
    return 0.5 * math.erfc(z / math.sqrt(2))

CITIES = ['CLUJ-NAPOCA', 'IAȘI', 'TIMIȘOARA']
out = {'min_candidati_per_scoala_an': MIN_N, 'ani': {}}

for year in YEARS:
    acc = per_year[year]
    groups = defaultdict(list)
    for code, vals in acc.items():
        if len(vals) >= MIN_N:
            groups[registry[code]].append(sum(vals) / len(vals))
    glist = [groups[c] for c in CITIES]
    sizes = [len(g) for g in glist]
    H, N, rank_sums = kruskal(glist)
    k = 3
    p = math.exp(-H / 2)
    eps2 = (H - k + 1) / (N - k)
    mean_ranks = [rs / n for rs, n in zip(rank_sums, sizes)]

    meds = {}
    for c in CITIES:
        v = sorted(groups[c])
        n = len(v)
        meds[c] = round(v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2, 3)

    all_v = [v for g in glist for v in g]
    cnt = Counter(all_v)
    tie_term = sum(t ** 3 - t for t in cnt.values()) / (12 * (N - 1))
    pairs = []
    for i in range(k):
        for j in range(i + 1, k):
            se = math.sqrt((N * (N + 1) / 12 - tie_term) * (1 / sizes[i] + 1 / sizes[j]))
            z = abs(mean_ranks[i] - mean_ranks[j]) / se
            praw = 2 * norm_sf(z)
            pairs.append([CITIES[i], CITIES[j], z, praw])
    pairs.sort(key=lambda t: t[3])
    prev = 0.0
    dunn = []
    for pi, (a, b, z, praw) in enumerate(pairs):
        padj = min(1.0, max(prev, (3 - pi) * praw))
        prev = padj
        dunn.append({'pereche': f'{a} vs {b}', 'z': round(z, 3), 'p_holm': round(padj, 5)})

    out['ani'][year] = {
        'n_scoli': dict(zip(CITIES, sizes)),
        'mediana_medii_scoli': meds,
        'rang_mediu': {c: round(mr, 1) for c, mr in zip(CITIES, mean_ranks)},
        'H': round(H, 3), 'p': float(f'{p:.4g}'), 'epsilon2': round(eps2, 4),
        'dunn_holm': dunn,
    }
    print(f"{year}: n={sizes} H={H:.2f} p={p:.4g} eps2={eps2:.3f} mediane={[meds[c] for c in CITIES]}")
    for dd in dunn:
        print('   ', dd['pereche'], 'z=', dd['z'], 'p_holm=', dd['p_holm'])

op = f"{BASE}\\kw_pe_ani.json"
with open(op, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print('saved', op)
