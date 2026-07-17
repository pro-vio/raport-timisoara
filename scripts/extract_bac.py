# Extrage candidații BAC (sesiunea de vară 2017-2025) din școlile din
# Timișoara / Cluj-Napoca / Iași într-un singur JSON normalizat:
#   [an, siiir, forma_inv, filiera, profil, promotia_curenta(0/1), status,
#    medie_pub|null, medie_calc|null]
# + meta cu contoare de validare per an.
#
# SURSELE. Toți anii se citesc din formate cu valori tipate, deci fără ambiguitate.
# Prima versiune a acestui script lua CSV-ul pentru 2017 și 2019 și pierdea 1,8% din
# rânduri: acele CSV-uri au separatorul zecimal VIRGULĂ ne-quotat („6,31" → două câmpuri),
# iar nota nu se poate reconstrui din format — „5,6 · 9" și „5 · 6,9" sunt amândouă
# valide și dau medii diferite. Pierderea era însă evitabilă: aceleași seturi de pe
# data.gov.ro conțin ȘI un ODS (2017), ȘI un XLSX (2019), unde valorile sunt numere.
# De aceea nu mai există cale de CSV aici.
#
# DE CE RECALCULĂM MEDIA. Coloana „Medie" din date există DOAR pentru candidații cu ≥5 la
# fiecare probă — verificat pe 2025: 84.464 din 107.961 (78%) au medie, iar cei 23.420
# (22%) care s-au prezentat și au picat o probă nu au niciuna; zero excepții. O mediană pe
# coloana publicată ar fi deci mediana SUPRAVIEȚUITORILOR, iar selecția e mai dură la
# liceele slabe — statistica lor s-ar calcula pe un subeșantion tot mai favorabil.
#
# FORMULA (verificată: reproduce exact media publicată în 100% din cazuri, în fiecare an):
# media aritmetică a notelor finale pe probe, TRUNCHIATĂ la 2 zecimale, nu rotunjită
# (regula BAC „două zecimale, fără rotunjire"; rotunjirea ar greși în 33% din cazuri).
# Nota finală pe probă = cea de la contestație acolo unde s-a contestat — necondiționat,
# verificat pe fiecare an 2017-2025 cu regula_contestatie.py. (2016 făcea excepție, cu
# regula veche „doar dacă |dif| >= 0,5", și e scos din analiză.)
# Media se calculează pentru status ∈ {Promovat, Nepromovat} — aceștia s-au prezentat la
# tot; Absent/Eliminat au probe lipsă (~3% din candidați).
#
# Particularități de format:
#  - 2017: ODS (30 MB) — citit în flux cu ods_reader.py; odfpy ar încărca tot în memorie.
#  - 2019, 2020: XLSX cu metadata de dimensiuni ruptă → read_only=False (lent, RAM mult).
#  - 2022: schemă proprie 74 col. (Cod SIIIR, FORMA_DE_INV, FILIERA/PROFIL, PROMOTIA_CURENTA,
#    STATUS_FINAL, MEDIA_FINALA unde -2 = fără medie). Coloanele de note sunt identice.
import sys, os, json, re
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter, defaultdict
import xlrd
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import ods_reader

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(HERE, '..', 'date', 'bac')
REG = os.path.join(HERE, '..', 'date', 'Unitati de invatamant acreditate  i autorizate.xls')

PROBE = ('EA', 'EB', 'EC', 'ED')
SIIIR_RE = re.compile(r'^\d{10}$')

def norm(s):
    if not isinstance(s, str):
        return s
    return (s.strip().upper().replace('Ş', 'Ș').replace('ş', 'ș')
            .replace('Ţ', 'Ț').replace('ţ', 'ț'))

def clean(v):
    return '' if v is None else str(v).strip()

def fara_diacritice(s):
    for a, b in (('ă', 'a'), ('â', 'a'), ('î', 'i'), ('ș', 's'), ('ş', 's'),
                 ('ț', 't'), ('ţ', 't')):
        s = s.replace(a, b)
    return s

def norm_cat(v):
    """Filiera/Profil normalizate: anii nu scriu la fel („Teoretica" vs „Teoretică")."""
    return fara_diacritice(clean(v).lower())

def trunc2(x):
    return int(x * 100 + 1e-6) / 100

def nota(v):
    """O notă/medie ca float în [1,10], altfel None (2022 pune -2 = fără medie)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        x = float(v)
    else:
        v = str(v).strip().replace(',', '.')
        if v == '':
            return None
        try:
            x = float(v)
        except ValueError:
            return None
    return x if 1.0 <= x <= 10.0 else None

def medie_din_note(note):
    note = [n for n in note if n is not None]
    if len(note) < 3:            # EA, EC, ED sunt minimul; EB doar la limbă maternă
        return None
    return trunc2(sum(note) / len(note))

def nota_finala(base, cont):
    return [c if c is not None else b for b, c in zip(base, cont)]

# --- registrul: SIIIR -> oraș (doar cele 3 orașe) ---
targets = {'TIMIȘOARA', 'CLUJ-NAPOCA', 'IAȘI'}
wb = xlrd.open_workbook(REG)
ws = wb.sheet_by_index(0)
idx = {h: i for i, h in enumerate(ws.row_values(0))}
oras_of, den_of = {}, {}
for r in range(1, ws.nrows):
    row = ws.row_values(r)
    if norm(row[idx['Localitate']]) in targets:
        code = str(row[idx['Cod']]).strip()
        if code.endswith('.0'):
            code = code[:-2]
        oras_of[code] = norm(row[idx['Localitate']])
        den_of[code] = row[idx['Denumire']]
print(f'registru: {len(oras_of)} coduri SIIIR în cele 3 orașe', flush=True)

records = []
meta = defaultdict(dict)

def proceseaza(an, header, rows, schema2022=False):
    """Comun tuturor surselor: primește antetul și un iterator de rânduri."""
    stat = Counter()
    h = {name: i for i, name in enumerate(header)}
    promo_tinta = f'{an-1}-{an}'
    if schema2022:
        i_siiir, i_forma = h['Cod SIIIR'], h['FORMA_DE_INV']
        i_filiera, i_profil = h['FILIERA'], h['PROFIL']
        i_promo, i_status, i_medie = h['PROMOTIA_CURENTA'], h['STATUS_FINAL'], h['MEDIA_FINALA']
    else:
        i_siiir, i_forma = h['Unitate (SIIIR)'], h['Forma de învățământ']
        i_filiera, i_profil = h['Fileira'], h['Profil']    # „Fileira" e typo în sursă
        i_promo, i_status, i_medie = h['Promoție'], h['STATUS'], h['Medie']
    n_h = len(header)
    for row in rows:
        row = list(row)
        if len(row) < n_h:            # ODS taie celulele goale de la coada rândului
            row += [None] * (n_h - len(row))
        if all(v is None or v == '' for v in row):
            continue
        siiir = clean(row[i_siiir])
        if siiir.endswith('.0'):
            siiir = siiir[:-2]
        if not SIIIR_RE.match(siiir):
            stat['drop_siiir_invalid'] += 1
            continue
        promo = (1 if clean(row[i_promo]).upper() == 'DA' else 0) if schema2022 \
            else (1 if clean(row[i_promo]) == promo_tinta else 0)
        status = clean(row[i_status])
        medie_pub = nota(row[i_medie])
        medie_calc = medie_pub
        if status in ('Promovat', 'Nepromovat'):
            base = [nota(row[h[f'NOTA_{p}']]) for p in PROBE]
            cont = [nota(row[h[f'NOTA_CONTESTATIE_{p}']])
                    if clean(row[h[f'CONTESTATIE_{p}']]) == 'Da' else None for p in PROBE]
            m = medie_din_note(nota_finala(base, cont))
            if medie_pub is None:
                medie_calc = m
            elif m is None or abs(m - medie_pub) > 1e-9:
                stat['VALIDARE_nepotrivire'] += 1     # formula trebuie să dea 100%
        stat['rânduri'] += 1
        if medie_pub is not None:
            stat['medie_oficială'] += 1
        elif medie_calc is not None:
            stat['medie_recuperată'] += 1
        else:
            stat['fără_medie'] += 1
        if siiir in oras_of:
            stat['în_3_orașe'] += 1
            if promo:
                stat['în_3_orașe_promoția_curentă'] += 1
            records.append([an, siiir, clean(row[i_forma]), norm_cat(row[i_filiera]),
                            norm_cat(row[i_profil]), promo, status, medie_pub, medie_calc])
    meta[an] = dict(stat)
    print(f'{an}: {dict(stat)}', flush=True)

def din_xlsx(an, read_only=True, schema2022=False):
    wb = load_workbook(os.path.join(BASE, f'bac_{an}_s1.xlsx'), read_only=read_only)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [clean(x) for x in next(rows)]
    proceseaza(an, header, rows, schema2022)
    wb.close()

def din_ods(an):
    it = ods_reader.randuri(os.path.join(BASE, f'bac_{an}_s1.ods'))
    header = [clean(x) for x in next(it)]
    proceseaza(an, header, it)

din_ods(2017)
for an in (2018, 2021, 2023, 2024, 2025):
    din_xlsx(an)
din_xlsx(2022, schema2022=True)
for an in (2019, 2020):                 # metadata ruptă: full mode, lent și cu RAM mult
    din_xlsx(an, read_only=False)

# --- rezumat + salvare ---
print('\nPe an (toată țara): medii oficiale, medii recuperate de noi, rânduri pierdute')
for an in sorted(meta):
    m = meta[an]
    rele = {k: v for k, v in m.items() if k.startswith('VALIDARE') or k.startswith('drop')}
    print(f'  {an}: rânduri={m.get("rânduri", 0):,} oficiale={m.get("medie_oficială", 0):,} '
          f'recuperate={m.get("medie_recuperată", 0):,} '
          f'{"| " + str(rele) if rele else "| curat"}')

I_PROMO, I_PUB, I_CALC = 5, 7, 8
pc = [r for r in records if r[I_PROMO]]
print(f'\ncandidați în cele 3 orașe: {len(records):,} '
      f'(promoția curentă: {len(pc):,} = {len(pc)/max(len(records),1):.1%})')
print('\npromoția curentă, pe an: acoperirea mediei')
for an in sorted({r[0] for r in pc}):
    an_r = [r for r in pc if r[0] == an]
    cu = sum(1 for r in an_r if r[I_CALC] is not None)
    pub = sum(1 for r in an_r if r[I_PUB] is not None)
    print(f'  {an}: n={len(an_r):,} | cu medie recalculată: {cu:,} ({cu/len(an_r):.1%}) '
          f'| cu medie publicată: {pub:,} ({pub/len(an_r):.1%}) | câștig: +{cu-pub:,}')

out = {
    'descriere': 'BAC sesiunea de vară, candidați din unitățile din Timișoara/Cluj-Napoca/Iași. '
                 'medie_calc = media oficială recalculată (trunc 2 zecimale) din notele finale '
                 'pe probe, disponibilă pentru toți cei prezenți la toate probele; medie_pub = '
                 'coloana din date, existentă doar pentru cei cu ≥5 la fiecare probă.',
    'câmpuri': ['an', 'siiir', 'forma_inv', 'filiera', 'profil', 'promotia_curenta',
                'status', 'medie_pub', 'medie_calc'],
    'meta_extracție': {str(k): v for k, v in sorted(meta.items())},
    'denumiri': {c: den_of[c] for c in sorted({r[1] for r in records})},
    'orase': {c: oras_of[c] for c in sorted({r[1] for r in records})},
    'candidati': records,
}
out_path = os.path.join(BASE, 'candidati_bac.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False)
print(f'\nsalvat: {out_path} ({os.path.getsize(out_path):,} B, {len(records):,} candidați)')
