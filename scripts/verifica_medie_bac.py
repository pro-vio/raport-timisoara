# Întrebare: pentru CINE există coloana „Medie" în datele BAC?
# Indiciu din parsarea 2016: minimul mediilor e exact 5,00 — deci media pare a fi
# publicată doar când toate probele sunt ≥5, adică e condiționată de promovarea
# fiecărei probe. Dacă e așa, mediana mediilor per liceu e o statistică pe un
# subeșantion selectat (cei care au trecut de toate probele), nu pe toți cei care
# au dat examenul.
#
# Verific pe un an XLSX curat (schemă standard, fără problema virgulei zecimale):
#  1. nota finală pe probă = NOTA_CONTESTATIE_x dacă s-a contestat, altfel NOTA_x;
#  2. există Medie ⟺ toate probele susținute sunt ≥5?
#  3. Medie == media aritmetică a notelor finale (rotunjită la 2 zecimale)?
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter
from openpyxl import load_workbook

AN = int(sys.argv[1]) if len(sys.argv) > 1 else 2025
HERE = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(HERE, '..', 'date', 'bac', f'bac_{AN}_s1.xlsx')

def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    v = str(v).strip().replace(',', '.')
    if v in ('', 'None', '-'):
        return None
    try:
        return float(v)
    except ValueError:
        return None

wb = load_workbook(path, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = [str(h).strip() if h is not None else '' for h in next(rows)]
h = {n: i for i, n in enumerate(header)}

PROBE = ['EA', 'EB', 'EC', 'ED']
c = Counter()
ex_fara_medie_dar_toate_peste5 = []
ex_cu_medie_dar_o_proba_sub5 = []
dif = []

for row in rows:
    if row is None or all(v is None for v in row):
        continue
    status = str(row[h['STATUS']]).strip() if row[h['STATUS']] is not None else ''
    note = []
    for p in PROBE:
        contest = str(row[h[f'CONTESTATIE_{p}']]).strip() if row[h[f'CONTESTATIE_{p}']] is not None else ''
        n_final = num(row[h[f'NOTA_CONTESTATIE_{p}']]) if contest == 'Da' else num(row[h[f'NOTA_{p}']])
        if n_final is not None:
            note.append(n_final)
    medie = num(row[h['Medie']])
    c[f'status={status}'] += 1
    c['total'] += 1
    if not note:
        c['fără nicio notă'] += 1
        continue
    toate_peste5 = all(n >= 5.0 for n in note)
    are_medie = medie is not None
    c[f'toate≥5={toate_peste5} & are_medie={are_medie}'] += 1
    if toate_peste5 and not are_medie and len(ex_fara_medie_dar_toate_peste5) < 5:
        ex_fara_medie_dar_toate_peste5.append((status, note))
    if not toate_peste5 and are_medie and len(ex_cu_medie_dar_o_proba_sub5) < 5:
        ex_cu_medie_dar_o_proba_sub5.append((status, note, medie))
    if are_medie:
        x = sum(note) / len(note)
        rotunjit = round(x, 2)
        trunchiat = int(x * 100 + 1e-6) / 100      # regula BAC: 2 zecimale, fără rotunjire
        c['ROTUNJIRE exact'] += (abs(rotunjit - medie) < 1e-9)
        c['TRUNCHIERE exact'] += (abs(trunchiat - medie) < 1e-9)
        c['niciuna'] += (abs(rotunjit - medie) >= 1e-9 and abs(trunchiat - medie) >= 1e-9)
        dif.append(round(abs(trunchiat - medie), 2))
        c[f'nr_probe_la_cei_cu_medie={len(note)}'] += 1

wb.close()

print(f'=== BAC {AN}, sesiunea de vară (toată țara) ===')
for k in sorted(c):
    print(f'  {k}: {c[k]:,}')
if dif:
    print(f'\nMedie recalculată vs publicată (n={len(dif):,}):')
    print(f'  identice (dif≤0,01): {sum(1 for d in dif if d <= 0.01):,} = '
          f'{sum(1 for d in dif if d <= 0.01)/len(dif):.1%}')
    print(f'  dif>0,05: {sum(1 for d in dif if d > 0.05):,}')
    print(f'  dif maximă: {max(dif)}')
print(f'\nex. toate≥5 dar FĂRĂ medie: {ex_fara_medie_dar_toate_peste5}')
print(f'ex. o probă <5 dar CU medie: {ex_cu_medie_dar_o_proba_sub5}')
